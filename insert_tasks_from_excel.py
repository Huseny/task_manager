#!/usr/bin/env python3
"""Convert Excel rows to Mindflow JSON and optionally post them to an API.

Examples:
	python insert_tasks_from_excel.py input.xlsx
	python insert_tasks_from_excel.py input.xlsx --endpoint http://127.0.0.1:8000/api/v1/mindflow-tasks --token YOUR_TOKEN --batch-no 4
"""

from __future__ import annotations

import argparse
import json
import os
from pathlib import Path
from typing import Any
from wsgiref import headers

import requests

try:
	from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - runtime dependency guard
	raise SystemExit(
		"openpyxl is required. Install it with: pip install openpyxl"
	) from exc


REQUIRED_COLUMNS = {
	"ID",
	"language",
	"types",
	"frontend_tech_stack",
	"backend_tech_stack",
	"database",
	"prompt",
}


def clean_value(value: Any) -> str:
	"""Return a trimmed string, normalizing common Excel number formatting."""

	if value is None:
		return ""
	if isinstance(value, float) and value.is_integer():
		return str(int(value))

	text = str(value).strip()
	if text.endswith(".0") and text[:-2].isdigit():
		return text[:-2]
	return text


def build_header_map(headers: list[Any]) -> dict[str, int]:
	return {
		str(header).strip(): index
		for index, header in enumerate(headers)
		if header is not None and str(header).strip()
	}


def row_to_record(row: tuple[Any, ...], header_map: dict[str, int]) -> dict[str, str] | None:
    if "ID" not in header_map:
        return None

    if not any(clean_value(cell) for cell in row):
        return None

    mindflow_id = clean_value(row[header_map["ID"]])
    prompt_text = (
        clean_value(row[header_map["prompt"]]) if "prompt" in header_map else ""
    )
    project_type = (
        clean_value(row[header_map["types"]]) if "types" in header_map else ""
    )

    model = "Claude Opus"

    frontend_stack = (
        clean_value(row[header_map["frontend_tech_stack"]])
        if "frontend_tech_stack" in header_map
        else ""
    )
    backend_stack = (
		clean_value(row[header_map["backend_tech_stack"]])
		if "backend_tech_stack" in header_map
		else ""
	)
    database = (
        clean_value(row[header_map["database"]]) if "database" in header_map else ""
    )

    if not any(
        [
            mindflow_id,
            prompt_text,
            project_type,
            frontend_stack,
            backend_stack,
            database,
        ]
    ):
        return None

    return {
        "mindflow_id": mindflow_id,
        "prompt_text": prompt_text,
        "project_type": project_type,
        "model": model,
        "frontend_tech_stack": frontend_stack,
        "backend_stack": backend_stack,
        "database": database,
    }


def convert_workbook(excel_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    records: list[dict[str, str]] = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
            print(f"Processing sheet: {sheet_name} with headers: {headers}")
        except StopIteration:
            continue

        header_map = build_header_map(headers)
        if not REQUIRED_COLUMNS.issubset(header_map.keys()):
            print(
                f"Skipping sheet '{sheet_name}' - missing required columns: "
                f"{REQUIRED_COLUMNS - set(header_map.keys())}"
            )
            continue

        for row in rows:
            record = row_to_record(row, header_map)
            if record is not None:
                records.append(record)

    return records


def load_bearer_token(cli_token: str | None, token_file: Path | None) -> str | None:
	if cli_token:
		return cli_token
	if token_file:
		return token_file.read_text(encoding="utf-8").strip()
	env_token = os.getenv("MINDFLOW_BEARER_TOKEN")
	return env_token.strip() if env_token else None


def send_tasks(
	tasks: list[dict[str, str]],
	endpoint_url: str,
	bearer_token: str,
	batch_no: str,
	timeout: float,
) -> tuple[int, int]:
    headers = {"Authorization": f"Bearer {bearer_token}"}
    success_count = 0
    failure_count = 0

    for i, task in enumerate(tasks):
        payload = dict(task)
        payload["batch_no"] = batch_no
        payload["task_id"] = f"w{batch_no}t{i+1}"
        task_identifier = payload.get("mindflow_id", "unknown")

        try:
            response = requests.post(
                endpoint_url, json=payload, headers=headers, timeout=timeout
            )
            if response.status_code in (200, 201):
                success_count += 1
                print(
                    f"Successfully sent mindflow_id: {task_identifier}, task_id: {payload['task_id']}"
                )
            else:
                failure_count += 1
                print(
                    f"Failed to send mindflow_id: {task_identifier}, task_id: {payload['task_id']} | "
                    f"Status: {response.status_code}"
                )
                print(f"Response: {response.text}")
        except requests.exceptions.RequestException as exc:
            failure_count += 1
            print(
                f"Connection error for mindflow_id {task_identifier}, task_id: {payload['task_id']}: {exc}"
            )

    return success_count, failure_count


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(
		description="Convert Excel rows into JSON records and optionally post them to an API."
	)
	parser.add_argument("excel_file", type=Path, help="Path to the input .xlsx file")
	parser.add_argument(
		"-o",
		"--output",
		type=Path,
		help="Path to the output .json file. Defaults to the Excel filename with .json suffix.",
	)
	parser.add_argument(
		"--skip-json-output",
		action="store_true",
		help="Do not write converted records to a JSON file.",
	)
	parser.add_argument(
		"--endpoint",
		type=str,
		help="API endpoint URL to POST each record to.",
	)
	parser.add_argument(
		"--token",
		type=str,
		help="Bearer token for API authentication.",
	)
	parser.add_argument(
		"--token-file",
		type=Path,
		help="Path to a file containing the Bearer token.",
	)
	parser.add_argument(
		"--batch-no",
		type=str,
		default="4",
		help="Batch number to inject into each task before sending. Default: 4",
	)
	parser.add_argument(
		"--timeout",
		type=float,
		default=30.0,
		help="HTTP timeout in seconds for each request. Default: 30",
	)
	return parser.parse_args()


def main() -> None:
	args = parse_args()
	excel_path = args.excel_file

	if not excel_path.exists():
		raise SystemExit(f"Input file not found: {excel_path}")

	records = convert_workbook(excel_path)

	if not args.skip_json_output:
		output_path = args.output or excel_path.with_suffix(".json")
		with output_path.open("w", encoding="utf-8") as handle:
			json.dump(records, handle, ensure_ascii=False, indent=2)
		print(f"Wrote {len(records)} records to {output_path}")

	if args.endpoint:
		token = load_bearer_token(args.token, args.token_file)
		if not token:
			raise SystemExit(
				"A bearer token is required when --endpoint is provided. "
				"Use --token, --token-file, or MINDFLOW_BEARER_TOKEN."
			)

		success_count, failure_count = send_tasks(
			tasks=records,
			endpoint_url=args.endpoint,
			bearer_token=token,
			batch_no=args.batch_no,
			timeout=args.timeout,
		)
		print(
			f"Finished sending tasks. Success: {success_count}, "
			f"Failed: {failure_count}, Total: {len(records)}"
		)


if __name__ == "__main__":
	main()
